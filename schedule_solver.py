import openpyxl
from datetime import datetime, timedelta
import re

# ========== 品类评级（从截图提取） ==========
CATEGORY_GRADES = {
    '编织': ('interest', 'B'),
    '短视频': ('interest', 'S'),
    '君合太极': ('health', 'S'),
    '懒人吃瘦': ('beauty', 'A'),
    '朗诵IP': ('interest', 'A'),
    '逆龄女神瑜伽': ('beauty', 'A'),
    '普拉提': ('beauty', 'A'),
    '摄影美学': ('interest', 'S'),
    '手机摄影': ('interest', 'S'),
    '睡眠调理': ('health', 'A'),
    '太极': ('health', 'A'),
    '五禽戏': ('health', 'A'),
    '一杰瑜伽': ('beauty', 'S'),
    '瑜伽': ('beauty', 'A'),
    '东方养正瑜伽': ('beauty', 'A'),
    '面部驻颜瑜伽': ('beauty', 'B'),
    '面部瑜伽': ('beauty', 'B'),
    '气血': ('health', 'B'),
    '健康营养': ('health', 'B'),
    '亚健康': ('health', 'B'),
    '体态塑形瑜伽': ('beauty', 'A'),
    '内养太极': ('health', 'A'),
    '古法居家姚国诚': ('health', 'B'),
    '固气活血': ('health', 'B'),
    '中医变美': ('beauty', 'A'),
}

GRADE_SCORE = {'S': 100, 'A': 70, 'B': 40, 'C': 20, None: 10}
LINE_LABEL = {'health': '健康线', 'beauty': '变美线', 'interest': '兴趣线'}
TARGET_EXPOSURE = {'S': 350000, 'A': 220000, 'B': 150000, 'C': 120000}

WEEK_DAYS = {2: '周二', 3: '周三', 4: '周四', 5: '周五', 6: '周六', 7: '周日'}


def parse_date_range(text):
    """从时间段文本解析开始和结束日期"""
    if not text:
        return None, None
    # 处理如 "2026年1月19日—2026年5月3日"
    m = re.search(r'(\d{4})年(\d{1,2})月\s*(\d{0,2})日?\s*[—~至]\s*(\d{4})年(\d{1,2})月\s*(\d{0,2})日?', text)
    if m:
        y1, mon1, d1, y2, mon2, d2 = m.groups()
        d1 = int(d1) if d1 else 1
        d2 = int(d2) if d2 else 1
        try:
            start = datetime(int(y1), int(mon1), d1)
            end = datetime(int(y2), int(mon2), d2)
            return start, end
        except:
            pass
    # 处理 "2023年1月" 这种只有开始日期的
    m = re.search(r'(\d{4})年(\d{1,2})月', text)
    if m:
        y, mon = m.groups()
        try:
            start = datetime(int(y), int(mon), 1)
            return start, start
        except:
            pass
    return None, None


def merge_date_ranges(ranges):
    """合并连续或重叠的时间段"""
    if not ranges:
        return []
    # 按开始日期排序
    sorted_ranges = sorted(ranges, key=lambda x: x[0])
    merged = [list(sorted_ranges[0])]
    for start, end, count in sorted_ranges[1:]:
        last_start, last_end, last_count = merged[-1]
        # 如果当前开始日期 <= 上一个结束日期 + 1天，则合并
        if start <= last_end + timedelta(days=1):
            merged[-1][1] = max(last_end, end)
            merged[-1][2] += count
        else:
            merged.append([start, end, count])
    return [(s, e, c) for s, e, c in merged]


def extract_live_info(schedule_wb):
    """从排期表提取所有直播信息"""
    ws = schedule_wb.active
    rows = list(ws.iter_rows(values_only=True))

    lives = []
    # 日期行
    date_row = rows[1]
    dates = {i: date_row[i] for i in range(2, 9)}  # 周一到周日，对应列 2-8

    # 早间晨练 (row 2) — 同单元格多行 = 一场联合直播
    morning_row = rows[2]
    for col in range(2, 9):
        cell = morning_row[col]
        if not cell:
            continue
        day_num = dates.get(col)
        if not day_num:
            continue
        # 一个单元格可能有多个直播，用换行分隔
        raw_lines = [l.strip() for l in str(cell).split('\n') if l.strip()]
        live_names = []
        for line in raw_lines:
            # 跳过非直播名称的行（如时间、备注）
            if re.match(r'\d{1,2}:\d{2}', line):
                continue
            if '已有单课id' in line or '复用' in line or '需剪辑' in line:
                continue
            if not line or line in ['早间', '晨练']:
                continue
            live_names.append(line)

        if not live_names:
            continue

        # 提取所有子品类的信息
        categories = [name.replace('晨练', '').strip() for name in live_names]
        grades = [CATEGORY_GRADES.get(cat, ('health', 'C'))[1] for cat in categories]
        lines_list = [CATEGORY_GRADES.get(cat, ('health', 'C'))[0] for cat in categories]

        # 联合直播以第一场为主
        primary_grade = grades[0]
        # 联合直播目标曝光：第一场目标 + 后续场次目标（经验值，落在30-40万区间）
        target = sum(TARGET_EXPOSURE.get(g, 150000) for g in grades)

        lives.append({
            'name': ' + '.join(live_names),
            'categories': categories,
            'category': categories[0],  # 主品类
            'lines': list(set(lines_list)),
            'line': lines_list[0],  # 主线
            'grade': primary_grade,
            'date': f'5/{int(day_num)}',
            'fullDate': f'2026-05-{int(day_num):02d}',
            'slot': 'morning',
            'type': 'real',
            'isJoint': len(live_names) > 1,
            'target': target,
        })

    # 晚IP专场 (row 7)
    evening_row = rows[7]
    for col in range(2, 9):
        cell = evening_row[col]
        if not cell:
            continue
        day_num = dates.get(col)
        if not day_num:
            continue
        lines = [l.strip() for l in str(cell).split('\n') if l.strip()]
        for line in lines:
            if re.match(r'\d{1,2}:\d{2}', line):
                continue
            if not line:
                continue
            # 推断品类
            if '短视频' in line:
                category = '短视频'
            elif '手机摄影' in line:
                category = '手机摄影'
            elif '朗诵' in line:
                category = '朗诵IP'
            else:
                category = line.split('-')[0].strip()
            line_info = CATEGORY_GRADES.get(category, ('interest', 'C'))
            lives.append({
                'name': line,
                'category': category,
                'line': line_info[0],
                'grade': line_info[1],
                'date': f'5/{int(day_num)}',
                'fullDate': f'2026-05-{int(day_num):02d}',
                'slot': 'evening',
                'type': 'real',
            })

    # 晚间IP专场 (row 13)
    evening2_row = rows[13]
    for col in range(2, 9):
        cell = evening2_row[col]
        if not cell:
            continue
        day_num = dates.get(col)
        if not day_num:
            continue
        lines = [l.strip() for l in str(cell).split('\n') if l.strip()]
        for line in lines:
            if re.match(r'\d{1,2}:\d{2}', line):
                continue
            if '数字人直播' in line:
                continue
            if not line:
                continue
            category = line.strip()
            line_info = CATEGORY_GRADES.get(category, ('health', 'C'))
            lives.append({
                'name': line,
                'category': category,
                'line': line_info[0],
                'grade': line_info[1],
                'date': f'5/{int(day_num)}',
                'fullDate': f'2026-05-{int(day_num):02d}',
                'slot': 'evening',
                'type': 'real',
            })

    # 伪直播复用 (row 19)
    fake_row = rows[19]
    fake_exposure_row = rows[21]
    for col in range(2, 9):
        cell = fake_row[col]
        if not cell:
            continue
        day_num = dates.get(col)
        if not day_num:
            continue
        lines = [l.strip() for l in str(cell).split('\n') if l.strip()]
        name = None
        for line in lines:
            if '伪直播复用' in line:
                continue
            if re.match(r'\d{1,2}:\d{2}', line):
                continue
            if '需剪辑' in line or '已有单课id' in line or '复用' in line:
                continue
            if line:
                name = line
                break
        if not name:
            continue
        # 推断品类
        if '编织' in name:
            category = '编织'
        elif '普拉提' in name:
            category = '普拉提'
        elif '摄影美学' in name:
            category = '摄影美学'
        elif '古法' in name:
            category = '古法居家姚国诚'
        else:
            category = name.split('-')[0].strip()
        line_info = CATEGORY_GRADES.get(category, ('health', 'C'))
        lives.append({
            'name': name,
            'category': category,
            'line': line_info[0],
            'grade': line_info[1],
            'date': f'5/{int(day_num)}',
            'fullDate': f'2026-05-{int(day_num):02d}',
            'slot': 'evening',
            'type': 'fake',
        })

    return lives


def extract_fake_excluded_audiences(schedule_wb):
    """提取伪直播要剔除的人群"""
    ws = schedule_wb.active
    rows = list(ws.iter_rows(values_only=True))
    date_row = rows[1]
    dates = {i: date_row[i] for i in range(2, 9)}

    excluded = []  # [(category, time_range_text, count)]
    fake_exposure_row = rows[21]

    for col in range(2, 9):
        cell = fake_exposure_row[col]
        if not cell:
            continue
        day_num = dates.get(col)
        if not day_num:
            continue
        date_str = f'5/{int(day_num)}'
        text = str(cell)
        # 提取 【上次宣发】或 【存量】后的人群信息
        # 格式如：2025年10月13日—2026年4月26日
        #         睡眠、五禽戏（71369）
        # 或：2023年1月—2026年4月26日
        #     亚健康（19558）

        # 先去掉"优先级：..."部分
        text = re.sub(r'优先级：[^\n]+', '', text)
        text = text.replace('【上次宣发】', '').replace('【存量】', '')

        # 用正则匹配 日期范围 + 品类（人数）
        # 先找到所有日期范围
        date_pattern = r'(\d{4})年(\d{1,2})月(?:\s*(\d{1,2})日)?\s*[—~至]\s*(\d{4})年(\d{1,2})月(?:\s*(\d{1,2})日)?'
        date_matches = list(re.finditer(date_pattern, text))

        for i, dm in enumerate(date_matches):
            y1, mon1, d1, y2, mon2, d2 = dm.groups()
            d1 = int(d1) if d1 else 1
            d2 = int(d2) if d2 else 1
            start_str = f'{y1}年{mon1}月{d1}日'
            end_str = f'{y2}年{mon2}月{d2}日'
            time_range = f'{start_str}—{end_str}'

            # 找到这个日期范围后面的品类和人数
            start_pos = dm.end()
            end_pos = date_matches[i + 1].start() if i + 1 < len(date_matches) else len(text)
            segment = text[start_pos:end_pos]

            # 提取品类和人数
            # 格式：品类（人数）
            item_pattern = r'([^（\n]+)（(\d+)）'
            for item_m in re.finditer(item_pattern, segment):
                cat = item_m.group(1).strip()
                count = int(item_m.group(2))
                excluded.append({
                    'category': cat,
                    'timeRange': time_range,
                    'count': count,
                    'date': date_str,
                })

    return excluded


def extract_excluded_from_confirmed(confirmed_wb):
    """从确认排期表提取伪直播 audience（上周已用，本周剔除）"""
    ws = confirmed_wb.active
    rows = list(ws.iter_rows(values_only=True))
    dates = {i: f'5/{rows[1][i]}' for i in range(2, 9)}

    excluded = []
    current_line = None
    for r in range(62, min(83, len(rows))):
        label_cell = rows[r][1] if len(rows[r]) > 1 else None
        if label_cell:
            label_text = str(label_cell).strip()
            if label_text in ['健康线', '变美线', '兴趣线']:
                current_line = label_text.replace('线', '')
                continue

        for col in range(2, 9):
            cell = rows[r][col]
            if cell is None:
                continue
            text = str(cell).strip()
            if text == '【存量】':
                continue
            if re.search(r'(\d{4})年', text) and '（' not in text and '(' not in text:
                continue
            m = re.search(r'(.+?)[（(](\d+)[)）]', text)
            if m:
                cat = m.group(1).strip()
                count = int(m.group(2))
                # 查找对应的时间段（向上扫描）
                time_range = None
                for rr in range(r - 1, 61, -1):
                    tr_cell = rows[rr][col]
                    if tr_cell and re.search(r'(\d{4})年', str(tr_cell)):
                        time_range = str(tr_cell).strip()
                        break
                excluded.append({
                    'date': dates[col],
                    'line': current_line,
                    'category': cat,
                    'count': count,
                    'timeRange': time_range,
                })
    return excluded


def extract_audience_segments(audience_wb):
    """从人数表提取 audience 段，合并同一品类的连续时间段"""
    ws = audience_wb.active
    rows = list(ws.iter_rows(values_only=True))

    # 找到表头行（第3行，索引2）
    # 表头：线, 品类, 品类目前总用户, 时间, 时间, 用户数, 时间, 用户数, 时间, 用户数, 时间, 用户数
    segments = []

    for row in rows[3:]:
        if not row[1]:
            continue
        line_name = row[0]  # 变美/健康/兴趣
        line = 'beauty' if line_name == '变美' else 'health' if line_name == '健康' else 'interest'
        category = str(row[1]).strip()
        # 提取所有时间段和人数
        # 列结构：品类(1), 总用户(2), 时间1(3), 时间2(4), 用户数1(5), 时间3(6), 用户数2(7), 时间4(8), 用户数3(9), 时间5(10), 用户数4(11)
        # 实际上表头是：线(0), 品类(1), 总用户(2), 时间(3), 时间(4), 用户数(5), 时间(6), 用户数(7), 时间(8), 用户数(9), 时间(10), 用户数(11)
        time_count_pairs = []
        # 时间段在列 3, 4, 6, 8, 10
        # 用户数在列 5, 7, 9, 11
        time_cols = [(3, 5), (4, 5), (6, 7), (8, 9), (10, 11)]
        for t_col, c_col in time_cols:
            time_text = row[t_col]
            count = row[c_col]
            if time_text and count is not None:
                count_val = int(count) if isinstance(count, (int, float)) else 0
                if count_val > 0:
                    start, end = parse_date_range(str(time_text))
                    if start and end:
                        time_count_pairs.append((start, end, count_val))

        # 合并连续时间段
        merged = merge_date_ranges(time_count_pairs)
        for start, end, count in merged:
            time_range_text = f'{start.year}年{start.month}月{start.day}日—{end.year}年{end.month}月{end.day}日'
            segments.append({
                'id': f'{category}_{start.strftime("%Y%m%d")}',
                'line': line,
                'category': category,
                'timeRange': time_range_text,
                'count': count,
                'status': 'available',
                'assigned_dates': [],
            })

    return segments


def is_same_category_family(a, b):
    """判断两个品类是否属于同一家族（严格相等）"""
    # 只有 normalizeCategory 后的标准名完全一致才算同族
    # 子串匹配和家族关键词匹配已移除
    return normalize_category(a) == normalize_category(b)


def normalize_category(name):
    """品类名规范化（简化版，与系统一致）"""
    s = name.strip()
    if not s:
        return ''
    # 直接匹配标准名
    if s in CATEGORY_GRADES:
        return s
    # 别名匹配
    aliases = {
        '睡眠': '睡眠调理',
        '五禽戏': '五禽戏',
        '瑜伽SA': '瑜伽',
        '普拉提SA': '普拉提',
        '太极SA': '太极',
        '手机摄影SA': '手机摄影',
        '手机摄影BCD': '手机摄影',
        '气血': '气血调理',
        '固气': '固气活血',
        '中医': '中医变美',
        '健康': '健康营养',
        '面部瑜伽': '面部驻颜瑜伽',
    }
    if s in aliases:
        return aliases[s]
    # 最长子串匹配
    best = ''
    for canonical in CATEGORY_GRADES.keys():
        if s in canonical and len(canonical) > len(best):
            best = canonical
    if best:
        return best
    for canonical in CATEGORY_GRADES.keys():
        if canonical in s and len(canonical) > len(best):
            best = canonical
    if best:
        return best
    return s


NEUTRAL_CATEGORIES = {'一杰瑜伽', '东方养正瑜伽'}


def get_live_day(date_str):
    """从 '5/11' 格式中提取日号"""
    return int(date_str.split('/')[1])


def auto_schedule(lives, audience_segments, excluded_audiences):
    """自动排期 - 联合直播模型、跨线、中性品类、3天复用"""
    segs = [dict(s) for s in audience_segments]

    # 1. 标记 excluded audiences 为全局已使用
    for ex in excluded_audiences:
        ex_cat = ex['category']
        ex_tr = ex.get('timeRange', '')
        ex_count = ex.get('count', 0)
        matched = False
        for seg in segs:
            if seg['category'] == ex_cat and seg['timeRange'] == ex_tr:
                seg['status'] = 'used'
                matched = True
                break
        if not matched:
            # 时间范围不完全匹配时，用品类+人数兜底
            for seg in segs:
                if seg['category'] == ex_cat and seg['count'] == ex_count:
                    seg['status'] = 'used'
                    break

    # 2. 伪直播不分配新 audience（只记录为空）
    fake_lives = [l for l in lives if l['type'] == 'fake']
    for live in fake_lives:
        live['assignedAudiences'] = []
        live['exposure'] = 0

    # 3. 按评级排序真直播
    scored = []
    for live in lives:
        if live['type'] == 'fake':
            continue
        score = GRADE_SCORE.get(live['grade'], 10)
        if live['slot'] == 'evening':
            score += 50
        else:
            score += 30
        scored.append((live, score))

    scored.sort(key=lambda x: -x[1])

    # 4. 分配 audience
    for live, score in scored:
        live['assignedAudiences'] = []
        live['exposure'] = 0
        target = live.get('target', TARGET_EXPOSURE.get(live['grade'], 150000))
        live_day = get_live_day(live['date'])

        # 确定可分配的线
        if live.get('isJoint'):
            allowed_lines = set(live.get('lines', [live['line']]))
        elif live['category'] in NEUTRAL_CATEGORIES and live['line'] == 'beauty':
            allowed_lines = {'beauty', 'health'}
        else:
            allowed_lines = {live['line']}

        # 确定需要排除的同族品类（normalize 后严格相等）
        if live.get('isJoint'):
            excluded_cats = set(live.get('categories', [live['category']]))
        else:
            excluded_cats = {live['category']}

        candidates = [
            s for s in segs
            if s['line'] in allowed_lines
            and s['status'] != 'used'
            and len(s.get('assigned_dates', [])) < 2
            and live_day not in s.get('assigned_dates', [])
            and all(abs(live_day - d) >= 3 for d in s.get('assigned_dates', []))
            and not any(is_same_category_family(s['category'], cat) for cat in excluded_cats)
        ]
        candidates.sort(key=lambda x: -x['count'])

        for seg in candidates:
            if live['exposure'] >= target:
                break
            live['assignedAudiences'].append(dict(seg))
            live['exposure'] += seg['count']
            seg.setdefault('assigned_dates', []).append(live_day)

    return [live for live, _ in scored] + fake_lives


def print_schedule_result(lives):
    """打印排期结果"""
    print("=" * 80)
    print("5月W2 (5.11-5.17) 排期结果")
    print("=" * 80)

    # 按日期和时段分组
    days = ['5/11', '5/12', '5/13', '5/14', '5/15', '5/16', '5/17']
    slots = ['morning', 'evening']

    for day in days:
        day_lives = [l for l in lives if l['date'] == day]
        if not day_lives:
            continue
        print(f"\n{'='*40}")
        date_num = int(day.split('/')[1])
        weekday = WEEK_DAYS.get(date_num, day)
        print(f"📅 {weekday} {day}")
        print(f"{'='*40}")

        for slot in slots:
            slot_lives = [l for l in day_lives if l['slot'] == slot]
            if not slot_lives:
                continue
            slot_name = '早间晨练' if slot == 'morning' else '晚间专场'
            print(f"\n  🕐 {slot_name}")
            for live in slot_lives:
                badge = f"[{live['grade']}级]" if live['grade'] else "[未评级]"
                line_badge = LINE_LABEL.get(live['line'], live['line'])
                fake_badge = " [伪直播]" if live['type'] == 'fake' else ""
                print(f"    📺 {live['name']} {badge} ({line_badge}){fake_badge}")
                target = live.get('target', TARGET_EXPOSURE.get(live['grade'], 150000))
                print(f"       目标曝光: {target:,}")
                print(f"       实际分配: {live['exposure']:,}")
                if live['assignedAudiences']:
                    for aud in live['assignedAudiences']:
                        print(f"       → {aud['category']} | {aud['timeRange']} | {aud['count']:,}人")
                else:
                    print(f"       → 未分配人群")

    # 汇总
    print(f"\n{'='*80}")
    print("📊 排期汇总")
    print(f"{'='*80}")
    total_exposure = sum(l['exposure'] for l in lives)
    print(f"总直播场次: {len(lives)}")
    print(f"总曝光量: {total_exposure:,}")
    for grade in ['S', 'A', 'B', 'C']:
        grade_lives = [l for l in lives if l['grade'] == grade]
        if grade_lives:
            print(f"  {grade}级: {len(grade_lives)}场, 总曝光 {sum(l['exposure'] for l in grade_lives):,}")


def export_to_excel(lives, filename='排期结果_5月W2.xlsx'):
    """导出排期结果到 Excel — 包含完整的进量时间段"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "排期结果"

    ws.append([
        '日期', '时段', '直播名称', '品类', '线级', '评级', '类型',
        '目标曝光', '实际曝光',
        '分配人群-品类', '分配人群-进量时间段', '分配人群-人数'
    ])

    for live in lives:
        audiences = live.get('assignedAudiences', [])
        if not audiences:
            ws.append([
                live['date'],
                '早间' if live['slot'] == 'morning' else '晚间',
                live['name'],
                live['category'],
                LINE_LABEL.get(live['line'], live['line']),
                live['grade'] or '未评级',
                '伪直播' if live['type'] == 'fake' else '正式直播',
                live.get('target', TARGET_EXPOSURE.get(live['grade'], 150000)),
                live['exposure'],
                '未分配', '—', 0
            ])
        else:
            # 每个 audience 段单独一行
            for i, a in enumerate(audiences):
                ws.append([
                    live['date'] if i == 0 else '',
                    ('早间' if live['slot'] == 'morning' else '晚间') if i == 0 else '',
                    live['name'] if i == 0 else '',
                    live['category'] if i == 0 else '',
                    LINE_LABEL.get(live['line'], live['line']) if i == 0 else '',
                    (live['grade'] or '未评级') if i == 0 else '',
                    ('伪直播' if live['type'] == 'fake' else '正式直播') if i == 0 else '',
                    live.get('target', TARGET_EXPOSURE.get(live['grade'], 150000)) if i == 0 else '',
                    live['exposure'] if i == 0 else '',
                    a['category'],
                    a['timeRange'],
                    a['count']
                ])

    # 调整列宽
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 36
    ws.column_dimensions['L'].width = 14

    wb.save(filename)
    print(f"\n✅ 排期结果已导出到: {filename}")


def main():
    print("📂 读取排期表...")
    schedule_wb = openpyxl.load_workbook('5月直播排期（带直播优先级和伪直播剔除人群）.xlsx')
    lives = extract_live_info(schedule_wb)
    # 过滤非本周排期：5/17 的摄影美学-段晓晖单人是上次复用记录，不是本周新排
    lives = [l for l in lives if not (l['date'] == '5/17' and '摄影美学' in l['name'] and l['type'] == 'fake')]
    print(f"   提取到 {len(lives)} 场直播")

    print("📂 读取人数表...")
    audience_wb = openpyxl.load_workbook('5月w2直播排期人数.xlsx')
    audience_segments = extract_audience_segments(audience_wb)
    print(f"   提取到 {len(audience_segments)} 个 audience 段")

    print("📂 提取伪直播剔除人群（从原始排期表）...")
    excluded_original = extract_fake_excluded_audiences(schedule_wb)
    print(f"   原始排期表剔除 {len(excluded_original)} 个人群段")

    print("📂 提取伪直播剔除人群（从确认排期表）...")
    confirmed_wb = openpyxl.load_workbook('5月确认排期.xlsx')
    excluded_confirmed = extract_excluded_from_confirmed(confirmed_wb)
    print(f"   确认排期表剔除 {len(excluded_confirmed)} 个人群段")

    # 合并两个来源的剔除人群
    excluded = excluded_original + excluded_confirmed
    print(f"   总计剔除 {len(excluded)} 个人群段")

    print("\n🔄 执行自动排期...")
    scheduled_lives = auto_schedule(lives, audience_segments, excluded)

    print_schedule_result(scheduled_lives)
    export_to_excel(scheduled_lives)


if __name__ == '__main__':
    main()
